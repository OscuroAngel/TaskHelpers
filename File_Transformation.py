#####################################################
# Script that helps to fix the broken line in the CSV file we receive from one of the suppliers.
# After exploring the file, I found the patterns of the broken line and the script to bring them back.
#####################################################

import pandas as pd
import openpyxl
import csv


# Define the input and output file paths
input_file = 'file_name.csv'
output_file = 'fixed_file_name.csv'


# Read the input CSV file and fix columns, replacing line breaks within double-quoted fields with spaces
def process_csv(input_file, output_file):
    try:
        with open(input_file, 'r', newline='', encoding='utf-16') as csvfile:
            # Use csv.reader to handle line breaks within fields
            reader = csv.reader(csvfile, delimiter=';', quotechar='"')

            fixed_rows = []
            for row in reader:
                # Ensure each row has 24 columns by appending empty strings
                while len(row) < 24:
                    row.append('')

                # Replace line breaks within fields with spaces
                fixed_row = [column.replace('\n', ' ').replace('\r', ' ') for column in row]
                fixed_rows.append(fixed_row)

        # Write the fixed rows to the output CSV file
        with open(output_file, 'w', newline='', encoding='utf-16') as csvfile_out:
            # Use csv.writer to handle writing rows with proper quoting and delimiters
            writer = csv.writer(csvfile_out, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerows(fixed_rows)

        print('CSV file processed and fixed successfully.')

    except Exception as e:
        print(f'Error: {e}')


# Call the function to process the CSV file
process_csv(input_file, output_file)
input('Press enter to continue: ')



# Define input and output file paths
input_file = 'fixed_file_name.csv'
output_file = 'output.xlsx'

print("Reading CSV file...")
df = pd.read_csv(input_file, delimiter=';', header=0, encoding='utf-16', low_memory=False)

print("CSV file read successfully.")
print("Processing data...")

print("Data processed successfully.")
def merge_cells(input_file):

    merged_row = []

    def merge_cells(df):
        for index, row in df.iterrows():
            if len(row[2]) > 0 and row[2].strip() != '':
                print(f'Data found in col 3')
                df.at[index, df.columns[1]] = str(row[df.columns[1]]) + str(row[2])

            if len(row[1]) > 0 and row[1].strip() != '':
                print(f'Data found in col 2')
                df.at[index, df.columns[0]] = str(row[df.columns[0]]) + str(row[1])

merge_cells(df)

# Save the modified DataFrame back to the original CSV file
df.to_csv(input_file, index=False)
print(f'The conversion is done. Changes saved to {input_file}')
input('Press to continue: ')
